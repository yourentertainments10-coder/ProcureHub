"""Consolidated Vendor Inventory workbook generator.

Builds ONE openpyxl workbook with one worksheet per vendor (named by the
vendor's existing Vendor Code, e.g. AR_CT), each containing that vendor's
latest ACTIVE inventory read straight from the database -- the database
stays the source of truth; this is a read-only projection of it.

This is the "output/synchronisation layer" for the temporary WhatsApp-Excel
path, deliberately separate from both the import logic and the Google Sheets
sync so either output can operate independently. The row/column shape mirrors
the Google Sheets sync (`backend/app/integrations/google_sheets/sync_service.py`)
so the temporary Excel and the future Sheets output stay consistent.

Pure business logic -- no FastAPI/WhatsApp/print()/input() here.
"""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from core.ingestion.column_detector import DESCRIPTION_HEADERS, normalise_header
from core.models import InventoryImport
from core.services import inventory_import_service, vendor_service
from core.time_utils import now_ist

WORKBOOK_FILENAME = "Vendor_Inventory.xlsx"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Same columns the Google Sheets sync writes, so the two outputs match.
_HEADERS = ["Vendor Part Number", "Description", "Quantity Available", "Price", "MRP", "Last Updated"]

# Excel worksheet titles: max 31 chars and none of : \ / ? * [ ]
_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_MAX_SHEET_TITLE = 31


def _sheet_title(vendor, used: set[str]) -> str:
    """Worksheet name from the vendor's permanent Vendor Code (falls back to a
    sanitised name / id only if a legacy vendor has no code). Sanitised to
    Excel's rules and de-duplicated within this workbook."""
    base = (vendor.vendor_code or vendor.name or f"V{vendor.id}").strip()
    base = _INVALID_SHEET_CHARS.sub("_", base)[:_MAX_SHEET_TITLE] or f"V{vendor.id}"
    title = base
    suffix = 2
    while title.lower() in used:
        # Keep within 31 chars while making it unique (extremely unlikely for codes).
        title = f"{base[: _MAX_SHEET_TITLE - 2]}_{suffix}"
        suffix += 1
    used.add(title.lower())
    return title


def _description(row) -> str:
    """Pull the description out of a DB inventory row. Prefers the resolved
    `Part.description` (canonical, if the import wrote one), then falls back
    to the raw source-file description column (matched by the same
    `DESCRIPTION_HEADERS` the detector uses -- handles MAHINDRA's "Name",
    BIJVASAN's "Part Descr", etc.)."""
    if getattr(row, "part", None) is not None and getattr(row.part, "description", None):
        return str(row.part.description)
    raw_data = getattr(row, "raw_data", None)
    if isinstance(raw_data, dict):
        for key, value in raw_data.items():
            if normalise_header(key) in DESCRIPTION_HEADERS:
                return str(value or "")
    return ""


def _append_streamed_sheet(
    workbook: openpyxl.Workbook, title: str, headers: list, rows: list
) -> None:
    """Add one worksheet to a WRITE-ONLY workbook: bold header, sized
    columns, then every row streamed straight to disk-backed storage.

    write-only mode is the whole point: the consolidated workbook now spans
    tens of thousands of rows (26k+ active inventory rows in production),
    and regular openpyxl keeps a Python cell object per cell -- one build
    peaked past Render's 512MB memory limit and took the whole service down
    (OOM restart, losing queued notifications). Streaming keeps the build to
    the size of ONE row at a time. Column widths must be set BEFORE rows are
    appended in write-only mode, so they're computed from the data first."""
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title=title)

    widths = [len(str(value)) for value in headers]
    for row in rows:
        for index, value in enumerate(row):
            length = len(str(value)) if value is not None else 0
            if index >= len(widths):
                widths.append(length)
            elif length > widths[index]:
                widths[index] = length
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 80) + 2

    from openpyxl.cell import WriteOnlyCell

    header_cells = []
    for value in headers:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = Font(bold=True)
        header_cells.append(cell)
    sheet.append(header_cells)
    for row in rows:
        sheet.append(row)


def build_workbook(session: Session) -> openpyxl.Workbook:
    """One workbook, one worksheet per vendor (by Vendor Code), each holding
    that vendor's latest active inventory from the DB. Includes ALL vendors in
    the database -- not only the one that just imported -- so the workbook
    always represents the current full inventory state. Built in WRITE-ONLY
    (streaming) mode -- see `_append_streamed_sheet` for why."""
    workbook = openpyxl.Workbook(write_only=True)

    # Business timezone is IST -- see core.time_utils (single source of truth).
    synced_at = now_ist().strftime("%Y-%m-%d %H:%M IST")
    used_titles: set[str] = set()
    sheet_count = 0

    vendors = vendor_service.list_vendors(session)
    for vendor in vendors:
        rows = inventory_import_service.get_active_inventory(vendor.id, session)
        if not rows:
            # A vendor with no CURRENT (active) inventory gets no worksheet at
            # all -- rather than an empty, obsolete tab that looks like real
            # data. The database is the source of truth, so when a vendor's
            # inventory stops being current its sheet simply disappears from
            # the next generated workbook.
            continue

        # EXACT COPY of the vendor's own file: original columns, original
        # order, original values -- nothing renamed or filtered (the
        # normalized columns remain internal, for comparison/allocation).
        raw_headers, raw_rows = inventory_import_service.get_active_raw_table(
            vendor.id, session
        )
        if not raw_headers:
            # Legacy rows without captured raw cells: normalized fallback.
            raw_headers = list(_HEADERS)
            raw_rows = [
                [
                    row.vendor_part_number,
                    _description(row),
                    str(row.quantity_available),
                    str(row.price) if row.price is not None else "",
                    str(row.mrp) if row.mrp is not None else "",
                    synced_at,
                ]
                for row in rows
            ]
        _append_streamed_sheet(
            workbook, _sheet_title(vendor, used_titles), raw_headers, raw_rows
        )
        sheet_count += 1

    # An empty workbook is invalid; guarantee at least one sheet. In practice
    # this path is only ever reached after a successful import (>=1 vendor).
    if sheet_count == 0:
        workbook.create_sheet(title="Vendor_Inventory")

    return workbook


def build_import_workbook(import_id: int, session: Session) -> openpyxl.Workbook | None:
    """ONE vendor, ONE import: that batch's rows as a single-worksheet
    workbook (same exact-copy rule as the consolidated one). Powers the
    per-row Download in Import History, so the Founder can pull just one
    vendor -- including an older, superseded batch -- without opening the
    full workbook. Returns None when that import stored no rows (e.g. a
    FAILED import); the caller reports that rather than sending an empty
    file."""
    import_row = session.get(InventoryImport, import_id)
    if import_row is None:
        return None
    rows = inventory_import_service.get_import_rows(import_id, session)
    if not rows:
        return None

    vendor = vendor_service.get_vendor(import_row.vendor_id, session)
    title = _sheet_title(vendor, set()) if vendor is not None else f"Import_{import_id}"

    raw_headers, raw_rows = inventory_import_service.get_import_raw_table(import_id, session)
    if not raw_headers:
        synced_at = now_ist().strftime("%Y-%m-%d %H:%M IST")
        raw_headers = list(_HEADERS)
        raw_rows = [
            [
                row.vendor_part_number,
                _description(row),
                str(row.quantity_available),
                str(row.price) if row.price is not None else "",
                str(row.mrp) if row.mrp is not None else "",
                synced_at,
            ]
            for row in rows
        ]
    # Streaming (write-only) build -- a single vendor batch reaches 9k+ rows
    # in production; see `_append_streamed_sheet` for the memory rationale.
    workbook = openpyxl.Workbook(write_only=True)
    _append_streamed_sheet(workbook, title, raw_headers, raw_rows)
    return workbook


def import_workbook_filename(import_row, vendor) -> str:
    """`ND_CT_stock_13_import_42.xlsx` -- vendor code + source file + import
    id, so several downloads never collide in the Downloads folder."""
    code = (getattr(vendor, "vendor_code", None) or getattr(vendor, "name", None) or "vendor")
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", str(getattr(import_row, "file_name", "") or "")).rsplit(
        ".", 1
    )[0]
    code = re.sub(r"[^A-Za-z0-9._-]+", "_", str(code))
    return f"{code}_{stem}_import_{import_row.id}.xlsx" if stem else f"{code}_import_{import_row.id}.xlsx"


def workbook_to_bytes(workbook: openpyxl.Workbook) -> bytes:
    """Serialise to bytes in memory -- no temp file to store or clean up."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
