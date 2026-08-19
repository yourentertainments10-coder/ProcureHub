"""Vendor Selection: lets the user (or the automatic rule engine, see
`core.services.rules.engine`) pick one or more vendors + quantities per
`CustomerOrderItem` from the Vendor Comparison report
(`vendor_comparison_service.py`, which deliberately never picks one itself).

A `VendorSelection` row is created or replaced per (order line, vendor)
pair -- each customer part can be split across several vendors when no
single vendor can cover the full requested quantity, but selecting/removing
one vendor's allocation never touches another vendor's allocation for the
same line, or another line's selections. Selections persist until explicitly
cleared/replaced (re-running Automatic Vendor Selection, or a manual
deselect) or a Purchase Order is generated from them (see
`purchase_order_generation_service.py`), which only flips their status.

`VendorSelection` doubles as the vendor-stock reservation ledger (see
`core.services.vendor_stock_service`): `upsert_selection` locks the vendor's
inventory row (`_find_active_vendor_offer`'s `.with_for_update()`) and checks
the requested quantity against what's left ACROSS EVERY customer order for
that vendor+part, not just this one order line -- this is what prevents the
same vendor stock from being allocated to two different customers at once.

Pure business logic -- no FastAPI/print()/input() here.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from core.ingestion.column_detector import decimal_to_string, normalise_part_number
from core.models import (
    CustomerOrderItem,
    InventoryImport,
    Part,
    PartAlias,
    Vendor,
    VendorInventory,
    VendorSelection,
)
from core.services import vendor_stock_service
from core.services.vendor_comparison_service import compare_vendors_for_order


def _matchable_part_numbers(normalized_part_number: str, session: Session) -> list[str]:
    """Every normalized number that denotes the SAME part as the input: the
    input itself, the part's canonical number, and all of its `PartAlias`
    numbers (any vendor). This is what lets an order line written against a
    vendor's ALTERNATE number ("Root Part Num" / OEM number) lock and consume
    the very same inventory row the primary number would -- one part, one
    reservation ledger, whichever number the customer wrote."""
    from core.services import part_link_service

    # A FOUNDER-DECLARED equivalence ("MF390300ML32 and MF390300ML are the
    # same part") widens the search before the Part graph is walked, so both
    # spellings reach each other's inventory rows and share ONE reservation
    # ledger -- ordering 20 under either number leaves the same remainder.
    seeds = {normalized_part_number}
    seeds.update(part_link_service.linked_numbers(seeds, session))

    part_ids = set(
        session.execute(
            select(Part.id).where(Part.canonical_part_number.in_(seeds))
        ).scalars()
    )
    part_ids.update(
        session.execute(
            select(PartAlias.part_id).where(PartAlias.normalized_part_number.in_(seeds))
        ).scalars()
    )
    if not part_ids:
        return list(seeds)
    numbers = set(seeds)
    numbers.update(
        session.execute(
            select(Part.canonical_part_number).where(Part.id.in_(part_ids))
        ).scalars()
    )
    numbers.update(
        session.execute(
            select(PartAlias.normalized_part_number).where(PartAlias.part_id.in_(part_ids))
        ).scalars()
    )
    # Numbers reached through the Part graph may themselves be declared
    # equivalent to others -- one more pass closes the set.
    numbers.update(part_link_service.linked_numbers(numbers, session))
    return list(numbers)


def _find_active_vendor_offer(
    vendor_id: int, normalized_part_number: str, session: Session
) -> VendorInventory | None:
    """`.with_for_update()` makes this the lock point for the allocation
    guard below: two concurrent calls trying to allocate against the same
    vendor+part serialize here (a real row lock on Postgres/production; on
    SQLite -- which has no row-level lock -- its own single-writer file lock
    still serializes concurrent writers, just at a coarser grain, which is
    fine for dev/tests). The lock is released on commit/rollback of the
    caller's transaction, right after `upsert_selection` finishes writing.

    Matching is alias-aware (`_matchable_part_numbers`): the offer is found
    whichever of the part's known numbers the order line used. `.limit(1)` +
    a stable order keep the single-row lock semantics unchanged."""
    return (
        session.execute(
            select(VendorInventory)
            .join(InventoryImport, VendorInventory.import_id == InventoryImport.id)
            .where(
                VendorInventory.vendor_id == vendor_id,
                VendorInventory.normalized_part_number.in_(
                    _matchable_part_numbers(normalized_part_number, session)
                ),
                InventoryImport.is_active.is_(True),
            )
            .order_by(VendorInventory.id)
            .limit(1)
            .with_for_update()
        )
        .scalars()
        .first()
    )


def _selections_for_item(
    order_item_id: int, session: Session, *, exclude_vendor_id: int | None = None
) -> list[VendorSelection]:
    statement = select(VendorSelection).where(
        VendorSelection.customer_order_item_id == order_item_id
    )
    if exclude_vendor_id is not None:
        statement = statement.where(VendorSelection.vendor_id != exclude_vendor_id)
    return list(session.execute(statement).scalars())


def list_selections_for_item(order_item_id: int, session: Session) -> list[VendorSelection]:
    """All of one order line's current vendor allocations (there can be
    several when a line is split across vendors)."""
    return _selections_for_item(order_item_id, session)


def live_allocatable_quantity(order_item_id: int, vendor_id: int, session: Session) -> Decimal:
    """The most this vendor can be allocated for this order line RIGHT NOW:
    min(live remaining stock, this line's still-unfulfilled quantity) -- the
    same two guards `upsert_selection` enforces, exposed read-only. Takes the
    same `.with_for_update()` row lock as `upsert_selection`, so a caller that
    reads this and then upserts inside one transaction acts on a stable
    figure. Used by the automatic engine to clamp an allocation computed from
    a stale availability snapshot down to what a concurrent order left behind,
    instead of skipping the vendor entirely. Returns 0 when nothing can be
    allocated (no active offer, unresolved part, or nothing left)."""
    order_item = session.get(CustomerOrderItem, order_item_id)
    if order_item is None:
        return Decimal(0)
    normalized = normalise_part_number(order_item.part_number_raw)
    offer = _find_active_vendor_offer(vendor_id, normalized, session)
    if offer is None or offer.part_id is None:
        return Decimal(0)
    remaining = vendor_stock_service.remaining_quantity(
        vendor_id, offer.part_id, offer.quantity_available, session,
        exclude_order_item_id=order_item_id,
    )
    others = _selections_for_item(order_item_id, session, exclude_vendor_id=vendor_id)
    unfulfilled = order_item.quantity_requested - sum(
        (other.quantity_selected for other in others), Decimal(0)
    )
    return max(min(remaining, unfulfilled), Decimal(0))


def upsert_selection(
    order_item_id: int, vendor_id: int, quantity_selected: Decimal, session: Session
) -> VendorSelection:
    """Create or replace this vendor's allocation for one order line. An
    order line may have allocations from several vendors at once (e.g. when
    no single vendor can cover the full requested quantity) -- re-selecting
    the SAME vendor for the SAME line overwrites just that vendor's own
    allocation; it never touches any other vendor's allocation for the same
    line, or any other line's selections.

    Raises `LookupError` if the order item or vendor doesn't exist, and
    `ValueError` if the vendor doesn't currently stock the part,
    `quantity_selected` exceeds what that vendor has available, or the sum
    of all vendors' allocations for this line would exceed the requested
    quantity.
    """
    order_item = session.get(CustomerOrderItem, order_item_id)
    if order_item is None:
        raise LookupError(f"Customer order item {order_item_id} not found.")

    vendor = session.get(Vendor, vendor_id)
    if vendor is None:
        raise LookupError(f"Vendor {vendor_id} not found.")

    if quantity_selected <= 0:
        raise ValueError("quantity_selected must be greater than 0.")

    normalized = normalise_part_number(order_item.part_number_raw)
    offer = _find_active_vendor_offer(vendor_id, normalized, session)
    if offer is None:
        raise ValueError(
            f"Vendor '{vendor.name}' has no active inventory for part "
            f"{order_item.part_number_raw!r}."
        )

    if offer.part_id is None:
        raise ValueError(
            f"Vendor '{vendor.name}'s inventory row for "
            f"{order_item.part_number_raw!r} has no resolved Part -- re-import "
            "that vendor's inventory before selecting it."
        )

    # The actual double-allocation guard: how much of this vendor's part is
    # left once every OTHER customer order's reservation is accounted for --
    # not just this vendor's raw imported quantity. `exclude_order_item_id`
    # leaves this line's own existing selection (if any) for this same
    # vendor out of the "already reserved" sum, so re-selecting the same
    # vendor for the same line is judged against "what's free excluding what
    # I already had," not double-counted against itself. Computed AFTER the
    # row lock above, so two concurrent calls for the same vendor+part never
    # both see the same stale remaining figure.
    remaining = vendor_stock_service.remaining_quantity(
        vendor_id, offer.part_id, offer.quantity_available, session,
        exclude_order_item_id=order_item_id,
    )
    if quantity_selected > remaining:
        raise ValueError(
            f"Vendor '{vendor.name}' only has {remaining} of "
            f"{order_item.part_number_raw!r} remaining (requested "
            f"{quantity_selected}) once other customer orders' allocations are "
            f"accounted for -- {offer.quantity_available} is the vendor's raw "
            "imported quantity."
        )

    other_selections = _selections_for_item(
        order_item_id, session, exclude_vendor_id=vendor_id
    )
    already_allocated = sum(
        (other.quantity_selected for other in other_selections), Decimal(0)
    )
    if already_allocated + quantity_selected > order_item.quantity_requested:
        raise ValueError(
            f"Allocating {quantity_selected} to '{vendor.name}' would bring the total "
            f"selected for {order_item.part_number_raw!r} to "
            f"{already_allocated + quantity_selected}, exceeding the requested quantity "
            f"of {order_item.quantity_requested}."
        )

    selection = session.execute(
        select(VendorSelection).where(
            VendorSelection.customer_order_item_id == order_item_id,
            VendorSelection.vendor_id == vendor_id,
        )
    ).scalar_one_or_none()

    if selection is None:
        selection = VendorSelection(
            customer_order_item_id=order_item_id, vendor_id=vendor_id
        )
        session.add(selection)

    selection.part_id = offer.part_id
    selection.vendor_part_number = offer.vendor_part_number
    selection.quantity_selected = quantity_selected

    session.flush()
    return selection


def clear_selections_for_item(order_item_id: int, session: Session) -> None:
    """Remove every vendor's allocation for one order line. Used by the
    automatic rule engine before applying a fresh set of allocations, so
    re-running (or switching strategies) replaces the previous automatic
    picks instead of leaving stale vendors from an earlier run alongside the
    new ones."""
    for selection in _selections_for_item(order_item_id, session):
        session.delete(selection)
    session.flush()


def remove_selection(order_item_id: int, vendor_id: int, session: Session) -> None:
    """Remove one vendor's allocation from an order line, leaving any other
    vendors' allocations for that same line untouched. A no-op if that
    vendor had no allocation for this line."""
    selection = session.execute(
        select(VendorSelection).where(
            VendorSelection.customer_order_item_id == order_item_id,
            VendorSelection.vendor_id == vendor_id,
        )
    ).scalar_one_or_none()
    if selection is None:
        return
    session.delete(selection)
    session.flush()


def list_selections_for_order(order_id: int, session: Session) -> list[VendorSelection]:
    return list(
        session.execute(
            select(VendorSelection)
            .join(
                CustomerOrderItem,
                VendorSelection.customer_order_item_id == CustomerOrderItem.id,
            )
            .where(CustomerOrderItem.customer_order_id == order_id)
            .order_by(CustomerOrderItem.row_number)
        ).scalars()
    )


@dataclass
class SelectionExportRow:
    customer_part_number: str
    requested_quantity: Decimal
    vendor_name: str
    vendor_part_number: str
    available_quantity: Decimal | None
    selected_quantity: Decimal | None
    status: str
    reason: str = ""


def list_selections_for_export(order_id: int, session: Session) -> list[SelectionExportRow]:
    """The vendor-allocation report. For each order line it emits either:

    - one row PER selected vendor (showing that vendor's available and the
      quantity taken from it, Status = Fulfilled / Partial), or
    - a single "Cannot Fulfill" row when the total available across ALL
      matching vendors is less than requested (showing total available and the
      shortage) -- never a partial allocation, or
    - a single "Not Selected" row if a line has stock but no selection yet.

    Availability comes from the Vendor Comparison (the DB is the source of
    truth); selections provide the per-vendor allocated quantities."""
    comparison = compare_vendors_for_order(order_id, session)

    item_order: list[int] = []
    part_by_item: dict[int, str] = {}
    requested_by_item: dict[int, Decimal] = {}
    total_available_by_item: dict[int, Decimal] = defaultdict(lambda: Decimal(0))
    vendor_available_by_item: dict[int, dict[int, Decimal]] = defaultdict(dict)
    for row in comparison.rows:
        if row.order_item_id is None:
            continue
        if row.order_item_id not in part_by_item:
            item_order.append(row.order_item_id)
            part_by_item[row.order_item_id] = row.customer_part_number
            requested_by_item[row.order_item_id] = row.requested_quantity
        if row.vendor_id is not None and row.vendor_available_quantity:
            total_available_by_item[row.order_item_id] += row.vendor_available_quantity
            vendor_available_by_item[row.order_item_id][row.vendor_id] = row.vendor_available_quantity

    selections_by_item: dict[int, list[VendorSelection]] = defaultdict(list)
    for selection in list_selections_for_order(order_id, session):
        selections_by_item[selection.customer_order_item_id].append(selection)

    rows: list[SelectionExportRow] = []
    for order_item_id in item_order:
        part = part_by_item[order_item_id]
        requested = requested_by_item[order_item_id]
        total_available = total_available_by_item.get(order_item_id, Decimal(0))
        selections = selections_by_item.get(order_item_id, [])

        if selections:
            selected_total = sum((s.quantity_selected for s in selections), Decimal(0))
            status = "Fulfilled" if selected_total >= requested else "Partial"
            for selection in selections:
                vendor = session.get(Vendor, selection.vendor_id)
                rows.append(
                    SelectionExportRow(
                        customer_part_number=part,
                        requested_quantity=requested,
                        vendor_name=vendor.name if vendor else str(selection.vendor_id),
                        vendor_part_number=selection.vendor_part_number,
                        available_quantity=vendor_available_by_item[order_item_id].get(selection.vendor_id),
                        selected_quantity=selection.quantity_selected,
                        status=status,
                    )
                )
        elif total_available < requested:
            shortage = requested - total_available
            rows.append(
                SelectionExportRow(
                    customer_part_number=part,
                    requested_quantity=requested,
                    vendor_name="-",
                    vendor_part_number="-",
                    available_quantity=total_available,
                    selected_quantity=None,
                    status="Cannot Fulfill",
                    reason=f"Insufficient vendor stock (short {decimal_to_string(shortage)})",
                )
            )
        else:
            rows.append(
                SelectionExportRow(
                    customer_part_number=part,
                    requested_quantity=requested,
                    vendor_name="-",
                    vendor_part_number="-",
                    available_quantity=total_available,
                    selected_quantity=None,
                    status="Not Selected",
                )
            )

    return rows


EXPORT_HEADERS = [
    "Customer Part Number",
    "Requested Qty",
    "Vendor",
    "Vendor Part Number",
    "Available Qty",
    "Selected Qty",
    "Status",
    "Reason",
]


def _export_row_cells(row: SelectionExportRow) -> list[str]:
    """One export row. `available_quantity`/`selected_quantity` are shown as
    actual numbers when present; a missing selected quantity (Cannot Fulfill /
    Not Selected with no allocation) is shown as "0", never left blank, so the
    Selected Qty column always carries an explicit value."""
    return [
        row.customer_part_number,
        decimal_to_string(row.requested_quantity),
        row.vendor_name,
        row.vendor_part_number,
        decimal_to_string(row.available_quantity) if row.available_quantity is not None else "",
        decimal_to_string(row.selected_quantity) if row.selected_quantity is not None else "0",
        row.status,
        row.reason,
    ]


def _fill_export_sheet(sheet, rows: list[SelectionExportRow], heading: str | None = None) -> None:
    if heading:
        # Whose order this is, INSIDE the sheet -- a worksheet tab caps at 31
        # characters, so a long customer name can only be read here.
        sheet.append([heading])
        sheet["A1"].font = Font(bold=True, size=12)
        sheet.append([])

    header_row_index = sheet.max_row + 1 if heading else 1
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[header_row_index]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(_export_row_cells(row))

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def to_export_workbook(rows: list[SelectionExportRow]) -> openpyxl.Workbook:
    """Vendor allocation sheet: "this customer part will be sourced from
    this vendor" -- no pricing, no vendor grouping, no PO numbering."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Selected Vendors"
    _fill_export_sheet(sheet, rows)
    return workbook


def to_batch_export_workbook(
    reports: list[tuple[str, list[SelectionExportRow]]],
    headings: dict[str, str] | None = None,
) -> openpyxl.Workbook:
    """ONE workbook holding every customer order's allocation report as its
    own worksheet -- the batched counterpart of `to_export_workbook`, used by
    the automatic allocation batch. (WhatsApp's Cloud API rejects ZIP uploads,
    so 'all reports together in one file' is delivered as a multi-sheet
    Excel, mirroring the Vendor_Inventory.xlsx one-sheet-per-vendor shape.)
    `reports` is a list of (worksheet title, export rows), one per order."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()
    for title, rows in reports:
        # Excel worksheet titles: max 31 chars, no : \ / ? * [ ] -- and
        # unique within the workbook.
        base = re.sub(r"[:\\/?*\[\]]", "_", (title or "Order").strip())[:31] or "Order"
        candidate, suffix = base, 2
        while candidate.lower() in used:
            candidate = f"{base[:28]}_{suffix}"
            suffix += 1
        used.add(candidate.lower())
        _fill_export_sheet(
            workbook.create_sheet(title=candidate),
            rows,
            heading=(headings or {}).get(title),
        )
    if not workbook.sheetnames:  # an empty workbook is invalid
        workbook.create_sheet(title="No Allocations")
    return workbook
