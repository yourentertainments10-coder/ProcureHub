"""Top-level fulfillment summary across all vendors: overall KPIs, top/worst
performing vendors, and the parts most frequently left pending.

Run after `delivery_import.py` (independently of the other reporting
scripts -- it computes everything itself):

    python summary_report.py

Writes `output/summary_report.xlsx`.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import func, select

from core.db import get_session
from core.ingestion.column_detector import decimal_to_string
from core.logging_setup import get_logger
from core.models import PurchaseOrder, PurchaseOrderItem
from core.services.gap_analysis_service import compute_gap_analysis
from core.services.vendor_performance_service import compute_vendor_summaries

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "summary_report.xlsx"

TOP_N = 10

logger = get_logger("summary_report")


def _bold_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def write_summary_report(kpis, top_vendors, worst_vendors, missing_parts) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()

    kpi_sheet = workbook.active
    kpi_sheet.title = "Summary"
    kpi_sheet.append(["Metric", "Value"])
    for label, value in kpis:
        kpi_sheet.append([label, value])
    _bold_header(kpi_sheet)
    _autosize_columns(kpi_sheet)

    top_sheet = workbook.create_sheet("Top Vendors")
    top_sheet.append(["Vendor", "Ordered", "Delivered", "Pending", "Accuracy %"])
    for vendor in top_vendors:
        top_sheet.append(
            [
                vendor.vendor_name,
                decimal_to_string(vendor.ordered_qty),
                decimal_to_string(vendor.delivered_qty),
                decimal_to_string(vendor.pending_qty),
                float(vendor.accuracy_pct),
            ]
        )
    _bold_header(top_sheet)
    _autosize_columns(top_sheet)

    worst_sheet = workbook.create_sheet("Worst Performing Vendors")
    worst_sheet.append(["Vendor", "Ordered", "Delivered", "Pending", "Accuracy %"])
    for vendor in worst_vendors:
        worst_sheet.append(
            [
                vendor.vendor_name,
                decimal_to_string(vendor.ordered_qty),
                decimal_to_string(vendor.delivered_qty),
                decimal_to_string(vendor.pending_qty),
                float(vendor.accuracy_pct),
            ]
        )
    _bold_header(worst_sheet)
    _autosize_columns(worst_sheet)

    missing_sheet = workbook.create_sheet("Most Frequently Missing Parts")
    missing_sheet.append(["Part Number", "Times Pending"])
    for part_number, count in missing_parts:
        missing_sheet.append([part_number, count])
    _bold_header(missing_sheet)
    _autosize_columns(missing_sheet)

    workbook.save(OUTPUT_FILE)


def main() -> None:
    logger.info("Building fulfillment summary report ...")

    with get_session() as session:
        gap_rows = compute_gap_analysis(session)
        vendor_summaries = compute_vendor_summaries(gap_rows)

        total_vendors = len(vendor_summaries)
        total_orders = session.execute(
            select(func.count()).select_from(PurchaseOrderItem)
        ).scalar_one()
        total_purchase_orders = session.execute(
            select(func.count()).select_from(PurchaseOrder)
        ).scalar_one()

    if not gap_rows:
        print("No purchase order items found. Run po_generator.py first.")
        return

    total_delivered = sum((row.delivered_qty for row in gap_rows), 0)
    total_pending = sum((row.pending_qty for row in gap_rows), 0)

    vendors_by_accuracy = sorted(
        vendor_summaries, key=lambda vendor: vendor.accuracy_pct, reverse=True
    )
    top_vendors = vendors_by_accuracy[:TOP_N]
    worst_vendors = list(reversed(vendors_by_accuracy))[:TOP_N]

    missing_part_counts = Counter(
        row.part_number for row in gap_rows if row.pending_qty > 0
    )
    missing_parts = missing_part_counts.most_common(TOP_N)

    kpis = [
        ("Total Vendors", total_vendors),
        ("Total Orders (PO line items)", total_orders),
        ("Total Purchase Orders", total_purchase_orders),
        ("Total Delivered", decimal_to_string(total_delivered)),
        ("Total Pending", decimal_to_string(total_pending)),
    ]

    write_summary_report(kpis, top_vendors, worst_vendors, missing_parts)

    print("=" * 70)
    print("FULFILLMENT SUMMARY")
    for label, value in kpis:
        print(f"{label:<32}: {value}")
    print("-" * 70)
    print("Top vendors (by accuracy):")
    for vendor in top_vendors:
        print(f"  {vendor.vendor_name:<20} {vendor.accuracy_pct}%")
    print("Worst performing vendors:")
    for vendor in worst_vendors:
        print(f"  {vendor.vendor_name:<20} {vendor.accuracy_pct}%")
    print("Most frequently missing parts:")
    for part_number, count in missing_parts:
        print(f"  Part {part_number:<10} pending on {count} PO line(s)")
    print("-" * 70)
    print(f"Report written to: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
