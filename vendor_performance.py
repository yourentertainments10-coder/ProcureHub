"""Vendor Performance Dashboard: aggregate ordered/delivered/pending
quantities and accuracy per vendor and per vendor+part, write
`output/vendor_dashboard.xlsx`, and render summary charts to `charts/`.

Run after `delivery_import.py` (independently of `gap_analysis.py` /
`alternative_vendor.py` -- it computes gap analysis itself):

    python vendor_performance.py

Vendor Accuracy % = Delivered Quantity / Ordered Quantity x 100.
"""

from __future__ import annotations

from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font

from core.db import get_session
from core.ingestion.column_detector import decimal_to_string
from core.logging_setup import get_logger
from core.services.vendor_performance_service import compute_all

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
CHARTS_DIR = BASE_DIR / "charts"
DASHBOARD_FILE = OUTPUT_DIR / "vendor_dashboard.xlsx"

logger = get_logger("vendor_performance")


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def _bold_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def write_dashboard(vendor_summaries, part_performance, pending_items) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Vendor Summary"
    summary_sheet.append(["Vendor", "Ordered", "Delivered", "Pending", "Accuracy %"])
    for vendor in vendor_summaries:
        summary_sheet.append(
            [
                vendor.vendor_name,
                decimal_to_string(vendor.ordered_qty),
                decimal_to_string(vendor.delivered_qty),
                decimal_to_string(vendor.pending_qty),
                float(vendor.accuracy_pct),
            ]
        )
    _bold_header(summary_sheet)
    _autosize_columns(summary_sheet)

    part_sheet = workbook.create_sheet("Part-wise Performance")
    part_sheet.append(["Vendor", "Part Number", "Ordered", "Delivered", "Accuracy %"])
    for part in part_performance:
        part_sheet.append(
            [
                part.vendor_name,
                part.part_number,
                decimal_to_string(part.ordered_qty),
                decimal_to_string(part.delivered_qty),
                float(part.accuracy_pct),
            ]
        )
    _bold_header(part_sheet)
    _autosize_columns(part_sheet)

    pending_sheet = workbook.create_sheet("Pending Items")
    pending_sheet.append(["Vendor", "Part", "Pending Qty"])
    for item in pending_items:
        pending_sheet.append(
            [item.vendor_name, item.part_number, decimal_to_string(item.pending_qty)]
        )
    _bold_header(pending_sheet)
    _autosize_columns(pending_sheet)

    workbook.save(DASHBOARD_FILE)


def render_charts(vendor_summaries) -> None:
    CHARTS_DIR.mkdir(parents=True, exist_ok=True)

    vendor_names = [vendor.vendor_name for vendor in vendor_summaries]

    # Bar chart: vendor accuracy %
    accuracy_values = [float(vendor.accuracy_pct) for vendor in vendor_summaries]
    figure, axis = plt.subplots(figsize=(8, 5))
    axis.bar(vendor_names, accuracy_values, color="#4C72B0")
    axis.set_title("Vendor Accuracy %")
    axis.set_ylabel("Accuracy %")
    axis.set_ylim(0, max(100, max(accuracy_values, default=0) + 10))
    axis.tick_params(axis="x", rotation=30)
    figure.tight_layout()
    figure.savefig(CHARTS_DIR / "vendor_accuracy.png")
    plt.close(figure)

    # Bar chart: ordered vs delivered, grouped per vendor
    ordered_values = [float(vendor.ordered_qty) for vendor in vendor_summaries]
    delivered_values = [float(vendor.delivered_qty) for vendor in vendor_summaries]
    x_positions = range(len(vendor_names))
    bar_width = 0.35

    figure, axis = plt.subplots(figsize=(8, 5))
    axis.bar(
        [x - bar_width / 2 for x in x_positions], ordered_values, bar_width, label="Ordered", color="#4C72B0"
    )
    axis.bar(
        [x + bar_width / 2 for x in x_positions], delivered_values, bar_width, label="Delivered", color="#DD8452"
    )
    axis.set_title("Ordered vs Delivered Quantity")
    axis.set_ylabel("Quantity")
    axis.set_xticks(list(x_positions))
    axis.set_xticklabels(vendor_names, rotation=30)
    axis.legend()
    figure.tight_layout()
    figure.savefig(CHARTS_DIR / "ordered_vs_delivered.png")
    plt.close(figure)

    # Pie chart: total fulfilled vs pending
    total_delivered = sum(float(vendor.delivered_qty) for vendor in vendor_summaries)
    total_pending = sum(float(vendor.pending_qty) for vendor in vendor_summaries)

    figure, axis = plt.subplots(figsize=(6, 6))
    if total_delivered + total_pending > 0:
        axis.pie(
            [total_delivered, total_pending],
            labels=["Fulfilled", "Pending"],
            autopct="%1.1f%%",
            colors=["#55A868", "#C44E52"],
        )
    axis.set_title("Fulfilled vs Pending")
    figure.tight_layout()
    figure.savefig(CHARTS_DIR / "pending_items.png")
    plt.close(figure)


def main() -> None:
    logger.info("Computing vendor performance ...")

    with get_session() as session:
        vendor_summaries, part_performance, pending_items = compute_all(session)

    if not vendor_summaries:
        print("No purchase order items found. Run po_generator.py first.")
        return

    write_dashboard(vendor_summaries, part_performance, pending_items)
    render_charts(vendor_summaries)

    print("=" * 70)
    print("VENDOR PERFORMANCE")
    print(f"{'Vendor':<20}{'Ordered':>10}{'Delivered':>12}{'Pending':>10}{'Accuracy%':>12}")
    for vendor in vendor_summaries:
        print(
            f"{vendor.vendor_name:<20}"
            f"{decimal_to_string(vendor.ordered_qty):>10}"
            f"{decimal_to_string(vendor.delivered_qty):>12}"
            f"{decimal_to_string(vendor.pending_qty):>10}"
            f"{str(vendor.accuracy_pct):>12}"
        )
    print("-" * 70)
    print(f"Dashboard written to : {DASHBOARD_FILE}")
    print(f"Charts written to    : {CHARTS_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
